"""
Report service — generate Excel/CSV reports and analytics data.
Uses openpyxl for Excel export.
"""
from datetime import datetime
import io
from app import mongo
from app.models.employee_model import EmployeeModel
from app.models.attendance_model import AttendanceModel
from app.models.leave_model import LeaveRequestModel
from app.models.payroll_model import PayrollModel
from app.models.performance_model import PerformanceReviewModel
from app.utils.helpers import get_month_date_range


def _employees():
    return mongo.db[EmployeeModel.COLLECTION]


def _attendance():
    return mongo.db[AttendanceModel.COLLECTION]


def _leave():
    return mongo.db[LeaveRequestModel.COLLECTION]


def _payroll():
    return mongo.db[PayrollModel.COLLECTION]


def _performance():
    return mongo.db[PerformanceReviewModel.COLLECTION]


# ── Analytics Data (JSON) ─────────────────────────────────────────────────────

def get_dashboard_stats():
    """Key HR metrics for the dashboard."""
    total_employees = _employees().count_documents({"status": "active"})
    today = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)

    present_today = _attendance().count_documents({
        "date": today,
        "status": {"$in": ["present", "work_from_home"]},
    })

    pending_leaves = _leave().count_documents({"status": "pending"})

    current_month = datetime.utcnow().month
    current_year = datetime.utcnow().year
    payroll_processed = _payroll().count_documents({
        "month": current_month,
        "year": current_year,
        "status": {"$in": ["processed", "paid"]},
    })

    # Department distribution
    dept_pipeline = [
        {"$match": {"status": "active"}},
        {"$group": {"_id": "$department", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
    ]
    dept_stats = list(_employees().aggregate(dept_pipeline))

    # Attendance rate (last 30 days)
    from datetime import timedelta
    thirty_days_ago = today - timedelta(days=30)
    att_pipeline = [
        {"$match": {"date": {"$gte": thirty_days_ago}}},
        {"$group": {
            "_id": "$status",
            "count": {"$sum": 1},
        }},
    ]
    att_stats = {row["_id"]: row["count"] for row in _attendance().aggregate(att_pipeline)}
    total_att = sum(att_stats.values()) or 1
    attendance_rate = round(
        (att_stats.get("present", 0) + att_stats.get("work_from_home", 0)) / total_att * 100, 1
    )

    # Gender distribution
    gender_pipeline = [
        {"$match": {"status": "active"}},
        {"$group": {"_id": "$gender", "count": {"$sum": 1}}},
    ]
    gender_stats = list(_employees().aggregate(gender_pipeline))

    # Recent joiners (last 30 days)
    recent_joiners = _employees().count_documents({
        "date_of_joining": {"$gte": thirty_days_ago},
        "status": "active",
    })

    return {
        "total_employees": total_employees,
        "present_today": present_today,
        "attendance_rate": attendance_rate,
        "pending_leave_requests": pending_leaves,
        "payroll_processed_this_month": payroll_processed,
        "recent_joiners": recent_joiners,
        "department_distribution": dept_stats,
        "gender_distribution": gender_stats,
        "attendance_breakdown": att_stats,
    }


def get_attendance_report(year, month, department=None):
    start, end = get_month_date_range(year, month)
    emp_query = {"status": "active"}
    if department:
        emp_query["department"] = department
    employees = list(_employees().find(emp_query))

    report = []
    for emp in employees:
        emp_id = str(emp["_id"])
        records = list(_attendance().find({
            "employee_id": emp_id,
            "date": {"$gte": start, "$lte": end},
        }))
        summary = {"present": 0, "absent": 0, "half_day": 0,
                   "work_from_home": 0, "on_leave": 0, "holiday": 0}
        total_hours = 0.0
        for r in records:
            s = r.get("status", "absent")
            if s in summary:
                summary[s] += 1
            total_hours += r.get("work_hours", 0)

        report.append({
            "emp_id": emp.get("emp_id"),
            "name": emp.get("full_name"),
            "department": emp.get("department"),
            "designation": emp.get("designation"),
            **summary,
            "total_work_hours": round(total_hours, 2),
        })
    return report


def get_leave_report(year, department=None, leave_type=None):
    emp_query = {"status": "active"}
    if department:
        emp_query["department"] = department
    employees = {str(e["_id"]): e for e in _employees().find(emp_query)}

    leave_query = {
        "start_date": {
            "$gte": datetime(year, 1, 1),
            "$lte": datetime(year, 12, 31, 23, 59, 59),
        }
    }
    if leave_type:
        leave_query["leave_type"] = leave_type

    leaves = list(_leave().find(leave_query))
    report = []
    for l in leaves:
        emp = employees.get(l.get("employee_id"), {})
        if not emp:
            continue
        report.append({
            "emp_id": emp.get("emp_id"),
            "name": emp.get("full_name"),
            "department": emp.get("department"),
            "leave_type": l.get("leave_type"),
            "start_date": l["start_date"].strftime("%Y-%m-%d") if l.get("start_date") else "",
            "end_date": l["end_date"].strftime("%Y-%m-%d") if l.get("end_date") else "",
            "days": l.get("days_requested"),
            "status": l.get("status"),
        })
    return report


def get_payroll_report(year, month, department=None):
    query = {"year": year, "month": month}
    if department:
        query["department"] = department
    docs = list(_payroll().find(query).sort("employee_name", 1))
    return [PayrollModel.to_dict(d) for d in docs]


def generate_excel_report(report_type, data):
    """
    Generate an in-memory Excel file from report data.
    Returns BytesIO buffer.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return None, "openpyxl is not installed. Run: pip install openpyxl"

    wb = openpyxl.Workbook()
    ws = wb.active

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a3c5e", end_color="1a3c5e", fill_type="solid")

    if not data:
        ws.title = "No Data"
        ws["A1"] = "No data available for this report."
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf, None

    ws.title = report_type.title()
    headers = list(data[0].keys())

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header.replace("_", " ").title())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(header, ""))

    # Auto-fit columns (approximate)
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, None
